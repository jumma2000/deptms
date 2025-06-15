from flask import Flask, render_template, request, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
import re
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from io import BytesIO
from flask import make_response
from openpyxl import Workbook
from io import BytesIO
from flask import make_response

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///customers.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'your_secret_key' # Change this to a strong, random key in production
db = SQLAlchemy(app)

# Models
class Customer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(100), nullable=False)
    phone_number = db.Column(db.String(20), unique=True, nullable=False)
    email = db.Column(db.String(100), nullable=True)
    balance = db.Column(db.Float, default=0.0)
    transactions = db.relationship('Transaction', backref='customer', lazy=True)

    def __repr__(self):
        return f"Customer('{self.full_name}', '{self.phone_number}', '{self.balance}')"

class Transaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    date = db.Column(db.String(10), nullable=False) # YYYY-MM-DD
    type = db.Column(db.String(10), nullable=False) # 'debt' or 'payment'
    amount = db.Column(db.Float, nullable=False)
    due_date = db.Column(db.String(10), nullable=True) # YYYY-MM-DD, only for 'debt' type
    notes = db.Column(db.Text, nullable=True)
    payment_method = db.Column(db.String(50), nullable=True) # Only for 'payment' type

    def __repr__(self):
        return f"Transaction('{self.customer_id}', '{self.date}', '{self.type}', '{self.amount}')"

# Routes
@app.route('/')
def index():
    search_query = request.args.get('search', '')
    if search_query:
        customers = Customer.query.filter(
            (Customer.full_name.ilike(f'%{search_query}%')) |
            (Customer.phone_number.ilike(f'%{search_query}%'))
        ).all()
    else:
        customers = Customer.query.all()
    return render_template('index.html', customers=customers, search_query=search_query)

@app.route('/add_customer', methods=['GET', 'POST'])
def add_customer():
    if request.method == 'POST':
        full_name = request.form['full_name']
        phone_number = request.form['phone_number']
        email = request.form['email']

        if not full_name or not phone_number:
            flash('الاسم الكامل ورقم الهاتف مطلوبان.', 'danger')
            return redirect(url_for('add_customer'))

        if Customer.query.filter_by(phone_number=phone_number).first():
            flash('رقم الهاتف هذا مسجل بالفعل.', 'danger')
            return redirect(url_for('add_customer'))

        new_customer = Customer(full_name=full_name, phone_number=phone_number, email=email)
        db.session.add(new_customer)
        db.session.commit()
        flash('تمت إضافة العميل بنجاح!', 'success')
        return redirect(url_for('index'))
    return render_template('add_customer.html')

@app.route('/edit_customer/<int:customer_id>', methods=['GET', 'POST'])
def edit_customer(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    if request.method == 'POST':
        customer.full_name = request.form['full_name']
        customer.phone_number = request.form['phone_number']
        customer.email = request.form['email']

        if not customer.full_name or not customer.phone_number:
            flash('الاسم الكامل ورقم الهاتف مطلوبان.', 'danger')
            return redirect(url_for('edit_customer', customer_id=customer.id))

        # Check for duplicate phone number, excluding the current customer
        existing_customer = Customer.query.filter(
            Customer.phone_number == customer.phone_number,
            Customer.id != customer.id
        ).first()
        if existing_customer:
            flash('رقم الهاتف هذا مسجل بالفعل لعميل آخر.', 'danger')
            return redirect(url_for('edit_customer', customer_id=customer.id))

        db.session.commit()
        flash('تم تحديث بيانات العميل بنجاح!', 'success')
        return redirect(url_for('customer_detail', customer_id=customer.id))
    return render_template('edit_customer.html', customer=customer)

@app.route('/customer/<int:customer_id>')
def customer_detail(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    transactions = Transaction.query.filter_by(customer_id=customer.id).order_by(Transaction.date.desc()).all()

    transactions_with_status = []
    for t in transactions:
        status = "N/A"
        is_due_soon = False
        is_overdue = False
        if t.type == 'debt':
            if t.due_date:
                due_date_obj = datetime.datetime.strptime(t.due_date, '%Y-%m-%d').date()
                if due_date_obj < datetime.date.today():
                    status = "متأخر"
                    is_overdue = True
                elif (due_date_obj - datetime.date.today()).days <= 7: # Due within 7 days
                    status = "مستحق قريباً"
                    is_due_soon = True
                else:
                    status = "مستحق"
            else:
                status = "مستحق" # If no due date, assume it's due
        elif t.type == 'payment':
            status = "مدفوع"
        
        transactions_with_status.append({
            'id': t.id,
            'customer_id': t.customer_id,
            'date': t.date,
            'type': t.type,
            'amount': t.amount,
            'due_date': t.due_date,
            'notes': t.notes,
            'payment_method': t.payment_method,
            'status': status,
            'is_due_soon': is_due_soon,
            'is_overdue': is_overdue
        })

    total_debt = sum(t.amount for t in transactions if t.type == 'debt')
    total_payments = sum(t.amount for t in transactions if t.type == 'payment')
    current_balance = total_debt - total_payments
    customer.balance = current_balance # Update customer balance in DB
    db.session.commit()

    return render_template('customer_detail.html', customer=customer, transactions=transactions_with_status,
                           total_debt=total_debt, total_payments=total_payments, current_balance=current_balance)

@app.route('/delete_customer/<int:customer_id>', methods=['POST'])
def delete_customer(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    try:
        # Delete all transactions associated with the customer
        Transaction.query.filter_by(customer_id=customer.id).delete()
        db.session.delete(customer)
        db.session.commit()
        flash('تم حذف العميل وجميع معاملاته بنجاح!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف العميل: {e}', 'danger')
    return redirect(url_for('index'))

@app.route('/add_transaction/<int:customer_id>', methods=['GET', 'POST'])
def add_transaction(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    if request.method == 'POST':
        transaction_type = request.form['type']
        date = request.form['date']
        amount = float(request.form['amount'])
        due_date = request.form.get('due_date') if transaction_type == 'debt' else None
        notes = request.form.get('notes')
        payment_method = request.form.get('payment_method')

        if not date or not amount:
            flash('التاريخ والمبلغ مطلوبان.', 'danger')
            return redirect(url_for('add_transaction', customer_id=customer.id))

        if amount <= 0:
            flash('المبلغ يجب أن يكون أكبر من صفر.', 'danger')
            return redirect(url_for('add_transaction', customer_id=customer.id))

        new_transaction = Transaction(
            customer_id=customer.id,
            date=date,
            type=transaction_type,
            amount=amount,
            due_date=due_date,
            notes=notes,
            payment_method=payment_method if transaction_type == 'payment' else None
        )
        db.session.add(new_transaction)
        db.session.commit()

        # Recalculate customer balance after transaction
        transactions = Transaction.query.filter_by(customer_id=customer.id).all()
        total_debt = sum(t.amount for t in transactions if t.type == 'debt')
        total_payments = sum(t.amount for t in transactions if t.type == 'payment')
        customer.balance = total_debt - total_payments
        db.session.commit()

        flash(f'تم تسجيل { "دين" if transaction_type == "debt" else "سداد" } بنجاح!', 'success')
        return redirect(url_for('customer_detail', customer_id=customer.id))
    # For GET request, pass today's date to the template
    return render_template('add_transaction.html', customer=customer, today_date=datetime.date.today().strftime('%Y-%m-%d'))

@app.route('/debt_report', methods=['GET'])
def debt_report():
    customers = Customer.query.all()
    selected_customer_id = request.args.get('customer_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status_filter = request.args.get('status_filter')

    query = Transaction.query.join(Customer).order_by(Transaction.date.desc())

    if selected_customer_id:
        query = query.filter(Transaction.customer_id == selected_customer_id)
    if start_date:
        query = query.filter(Transaction.date >= start_date)
    if end_date:
        query = query.filter(Transaction.date <= end_date)

    all_transactions = query.all()

    report_data = []
    total_debt = 0.0
    total_payments = 0.0
    current_balance = 0.0

    for transaction in all_transactions:
        status = "N/A"
        if transaction.type == 'debt':
            if transaction.due_date:
                due_date_obj = datetime.datetime.strptime(transaction.due_date, '%Y-%m-%d').date()
                if due_date_obj < datetime.date.today():
                    status = "متأخر"
                else:
                    status = "مستحق"
            else:
                status = "مستحق" # If no due date, assume it's due
            total_debt += transaction.amount
        elif transaction.type == 'payment':
            status = "مدفوع"
            total_payments += transaction.amount
        
        # Apply status filter
        if status_filter and status_filter != status:
            continue

        report_data.append({
            'customer_name': transaction.customer.full_name,
            'transaction_date': transaction.date,
            'type': 'دين' if transaction.type == 'debt' else 'سداد',
            'amount': transaction.amount,
            'due_date': transaction.due_date if transaction.type == 'debt' else 'N/A',
            'status': status,
            'payment_method': transaction.payment_method if transaction.type == 'payment' else 'N/A',
            'notes': transaction.notes
        })
    
    current_balance = total_debt - total_payments

    return render_template('debt_report.html',
                           customers=customers,
                           report_data=report_data,
                           selected_customer_id=selected_customer_id,
                           start_date=start_date,
                           end_date=end_date,
                           status_filter=status_filter,
                           total_debt=total_debt,
                           total_payments=total_payments,
                           current_balance=current_balance)


@app.route('/send_whatsapp/<int:customer_id>', methods=['GET', 'POST'])
def send_whatsapp(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    default_message_template = "مرحباً {customer_name} : الدين الحالي هو {current_balance} د.ل.\nنرجو منكم تسوية المبلغ في أقرب وقت ممكن."

    if request.method == 'POST':
        message_template = request.form.get('message_template', default_message_template)
        
        # Replace placeholders
        message = message_template.replace('{customer_name}', customer.full_name)
        message = message.replace('{current_balance}', f"{customer.balance:.2f}")

        # Encode message for URL
        import urllib.parse
        encoded_message = urllib.parse.quote(message)
        
        # Construct WhatsApp Web URL
        whatsapp_url = f"https://wa.me/{customer.phone_number}?text={encoded_message}"
        
        flash('تم إنشاء رابط واتساب. سيتم فتح WhatsApp Web في نافذة جديدة.', 'info')
        return redirect(whatsapp_url)

    # For GET request, render the template with default message
    message_preview = default_message_template.replace('{customer_name}', customer.full_name)
    message_preview = message_preview.replace('{current_balance}', f"{customer.balance:.2f}")

    return render_template('send_whatsapp.html', customer=customer, 
                           default_message_template=default_message_template,
                           message_preview=message_preview)

@app.route('/generate_debt_report_excel', methods=['GET'])
def generate_debt_report_excel():
    selected_customer_id = request.args.get('customer_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status_filter = request.args.get('status_filter')

    query = Transaction.query.join(Customer).order_by(Transaction.date.desc())

    if selected_customer_id:
        query = query.filter(Transaction.customer_id == selected_customer_id)
    if start_date:
        query = query.filter(Transaction.date >= start_date)
    if end_date:
        query = query.filter(Transaction.date <= end_date)

    all_transactions = query.all()

    report_data = []
    total_debt = 0.0
    total_payments = 0.0

    for transaction in all_transactions:
        status = "N/A"
        if transaction.type == 'debt':
            if transaction.due_date:
                due_date_obj = datetime.datetime.strptime(transaction.due_date, '%Y-%m-%d').date()
                if due_date_obj < datetime.date.today():
                    status = "متأخر"
                else:
                    status = "مستحق"
            else:
                status = "مستحق"
            total_debt += transaction.amount
        elif transaction.type == 'payment':
            status = "مدفوع"
            total_payments += transaction.amount
        
        if status_filter and status_filter != status:
            continue

        report_data.append([
            transaction.customer.full_name,
            transaction.date,
            'دين' if transaction.type == 'debt' else 'سداد',
            transaction.amount,
            transaction.due_date if transaction.type == 'debt' else 'N/A',
            status,
            transaction.payment_method if transaction.type == 'payment' else 'N/A',
            transaction.notes if transaction.notes else ''
        ])
    
    current_balance = total_debt - total_payments

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير ديون العملاء"

    # Set sheet to Right-to-Left
    ws.sheet_view.rightToLeft = True

    # Define styles
    from openpyxl.styles import Font, Border, Side, Alignment

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=16)
    default_font = Font(size=16)
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    center_aligned_text = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = ["العميل", "تاريخ المعاملة", "النوع", "المبلغ (د.ل)", "تاريخ الاستحقاق", "الحالة", "طريقة الدفع", "ملاحظات"]
    ws.append(headers)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_aligned_text

    # Data
    for row_idx, row_data in enumerate(report_data, 2): # Start from row 2 for data
        ws.append(row_data)
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = default_font
            cell.border = thin_border
            cell.alignment = center_aligned_text

    # Add thick border after data table
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=len(report_data) + 1, column=col_idx).border = thick_border

    # Summary
    ws.append([]) # Empty row for spacing
    
    summary_start_row = len(report_data) + 3 # Adjust row for summary
    
    ws.cell(row=summary_start_row, column=1, value="إجمالي الديون:").font = Font(bold=True, size=16)
    ws.cell(row=summary_start_row, column=2, value=total_debt).font = default_font
    ws.cell(row=summary_start_row, column=2).alignment = center_aligned_text
    ws.cell(row=summary_start_row + 1, column=1, value="إجمالي المدفوعات:").font = Font(bold=True, size=16)
    ws.cell(row=summary_start_row + 1, column=2, value=total_payments).font = default_font
    ws.cell(row=summary_start_row + 1, column=2).alignment = center_aligned_text
    ws.cell(row=summary_start_row + 2, column=1, value="الرصيد الحالي (صافي الدين):").font = Font(bold=True, size=16)
    ws.cell(row=summary_start_row + 2, column=2, value=current_balance).font = default_font
    ws.cell(row=summary_start_row + 2, column=2).alignment = center_aligned_text

    # Apply borders to summary
    for r_offset in range(3):
        for c_offset in range(2):
            ws.cell(row=summary_start_row + r_offset, column=1 + c_offset).border = thin_border

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter # Get the column letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2 # Add a little padding
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO object
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=debt_report.xlsx'
    return response

if __name__ == '__main__':
    with app.app_context():
        # db.drop_all() # Drop all tables - Commented out to prevent data loss on restart
        db.create_all() # Create tables based on models
    app.run(debug=True)